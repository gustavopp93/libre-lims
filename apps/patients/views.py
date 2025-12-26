import logging

from django.contrib import messages
from django.contrib.auth import authenticate, logout
from django.contrib.auth import login as auth_login
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import models
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse_lazy
from django.views import View
from django.views.decorators.http import require_GET
from django.views.generic import CreateView, FormView, ListView, RedirectView, TemplateView, UpdateView

from apps.patients.forms import LeadSourceForm, LoginForm, PatientForm, PatientUpdateForm
from apps.patients.models import LeadSource, Patient

logger = logging.getLogger(__name__)


class LoginView(FormView):
    template_name = "login.html"
    form_class = LoginForm
    success_url = reverse_lazy("dashboard")

    def form_valid(self, form):
        username = form.cleaned_data["username"]
        password = form.cleaned_data["password"]

        user = authenticate(self.request, username=username, password=password)

        if user is not None:
            auth_login(self.request, user)
            messages.success(self.request, "Inicio de sesión exitoso")
            return super().form_valid(form)
        else:
            messages.error(self.request, "Usuario o contraseña incorrectos")
            return self.form_invalid(form)


class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = "dashboard.html"
    login_url = reverse_lazy("login")


class LogoutView(RedirectView):
    permanent = False
    url = reverse_lazy("login")

    def get(self, request, *args, **kwargs):
        logout(request)
        messages.success(request, "Sesión cerrada exitosamente")
        return super().get(request, *args, **kwargs)


class PatientsListView(LoginRequiredMixin, ListView):
    model = Patient
    template_name = "patients/patient_list.html"
    context_object_name = "patients"
    paginate_by = 20
    login_url = reverse_lazy("login")

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filtrar por tipo de documento
        document_type = self.request.GET.get("document_type")
        if document_type:
            queryset = queryset.filter(document_type=document_type)

        # Filtrar por número de documento
        document_number = self.request.GET.get("document_number")
        if document_number:
            queryset = queryset.filter(document_number__icontains=document_number)

        return queryset.order_by("-created_at")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["document_type"] = self.request.GET.get("document_type", "")
        context["document_number"] = self.request.GET.get("document_number", "")
        context["breadcrumbs"] = [
            {"name": "Pacientes", "url": None},
        ]
        return context


class CreatePatientView(LoginRequiredMixin, CreateView):
    model = Patient
    form_class = PatientForm
    template_name = "patients/patient_create.html"
    login_url = reverse_lazy("login")

    def get_success_url(self):
        next_url = self.request.GET.get("next")
        if next_url:
            return next_url
        return reverse_lazy("patients_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["next_url"] = self.request.GET.get("next", "")
        context["breadcrumbs"] = [
            {"name": "Pacientes", "url": reverse_lazy("patients_list")},
            {"name": "Crear Paciente", "url": None},
        ]
        return context

    def form_valid(self, form):
        messages.success(self.request, "Paciente creado exitosamente")
        return super().form_valid(form)


class UpdatePatientView(LoginRequiredMixin, UpdateView):
    model = Patient
    form_class = PatientUpdateForm
    template_name = "patients/patient_update.html"
    success_url = reverse_lazy("patients_list")
    login_url = reverse_lazy("login")
    context_object_name = "patient"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["breadcrumbs"] = [
            {"name": "Pacientes", "url": reverse_lazy("patients_list")},
            {"name": f"Editar: {self.object.first_name} {self.object.last_name}", "url": None},
        ]
        return context

    def form_valid(self, form):
        messages.success(self.request, "Paciente actualizado exitosamente")
        return super().form_valid(form)


class PatientsDownloadTemplateView(LoginRequiredMixin, View):
    """Download template Excel for patient import"""

    login_url = reverse_lazy("login")

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plantilla Pacientes"

        # Headers: Apellidos, Nombres, Tipo Documento, Número Documento, Fecha Nacimiento, Sexo, Teléfono
        headers = ["Apellidos", "Nombres", "Tipo Documento", "Número Documento", "Fecha Nacimiento", "Sexo", "Teléfono"]
        ws.append(headers)

        # Estilizar headers (mismo estilo que pricing)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Fila de ejemplo
        ws.append(["García López", "Juan Carlos", "DNI", "12345678", "17/10/1990", "M", "987654321"])

        # Ajustar anchos de columna
        ws.column_dimensions["A"].width = 20  # Apellidos
        ws.column_dimensions["B"].width = 20  # Nombres
        ws.column_dimensions["C"].width = 15  # Tipo Documento
        ws.column_dimensions["D"].width = 15  # Número Documento
        ws.column_dimensions["E"].width = 18  # Fecha Nacimiento
        ws.column_dimensions["F"].width = 8  # Sexo
        ws.column_dimensions["G"].width = 15  # Teléfono

        # Create response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="plantilla_pacientes.xlsx"'

        wb.save(response)
        return response


class PatientsUploadView(LoginRequiredMixin, View):
    """Upload patients from Excel file"""

    login_url = reverse_lazy("login")
    template_name = "patients/patients_upload.html"

    def get(self, request):
        context = {
            "breadcrumbs": [
                {"name": "Pacientes", "url": reverse_lazy("patients_list")},
                {"name": "Importar Pacientes", "url": None},
            ],
        }
        return render(request, self.template_name, context)

    def post(self, request):
        if "file" not in request.FILES:
            messages.error(request, "No se seleccionó ningún archivo")
            return redirect("patients_list")

        excel_file = request.FILES["file"]

        try:
            from datetime import datetime

            import openpyxl

            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            created_count = 0
            skipped_count = 0
            error_count = 0

            # Mapeos
            document_type_map = {
                "DNI": Patient.DocumentType.DNI,
                "C.E": Patient.DocumentType.CE,
                "PAS": Patient.DocumentType.PASAPORTE,
            }

            sex_map = {"F": Patient.Sex.FEMALE, "M": Patient.Sex.MALE}

            # Iterar desde fila 2 (skip header)
            for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Columnas: Apellidos, Nombres, Tipo Documento, Número Documento, Fecha Nacimiento, Sexo, Teléfono
                last_name = row[0]
                first_name = row[1]
                document_type_excel = row[2]
                document_number = row[3]
                birthdate_str = row[4]
                sex_excel = row[5]
                phone_number = row[6]

                # Skip si faltan datos críticos
                if not all(
                    [
                        last_name,
                        first_name,
                        document_type_excel,
                        document_number,
                        birthdate_str,
                        sex_excel,
                        phone_number,
                    ]
                ):
                    missing_fields = []
                    if not last_name:
                        missing_fields.append("Apellidos")
                    if not first_name:
                        missing_fields.append("Nombres")
                    if not document_type_excel:
                        missing_fields.append("Tipo Documento")
                    if not document_number:
                        missing_fields.append("Número Documento")
                    if not birthdate_str:
                        missing_fields.append("Fecha Nacimiento")
                    if not sex_excel:
                        missing_fields.append("Sexo")
                    if not phone_number:
                        missing_fields.append("Teléfono")

                    logger.warning(
                        f"Fila {row_number}: Datos incompletos - Faltan campos: {', '.join(missing_fields)} "
                        f"[{last_name or 'N/A'}, {first_name or 'N/A'}, {document_type_excel or 'N/A'} {document_number or 'N/A'}]"
                    )
                    error_count += 1
                    continue

                try:
                    # Mapear tipo documento
                    document_type = document_type_map.get(document_type_excel)
                    if not document_type:
                        logger.warning(
                            f"Fila {row_number}: Tipo de documento inválido '{document_type_excel}' "
                            f"[{last_name}, {first_name}, {document_number}] - Valores permitidos: DNI, C.E, PAS"
                        )
                        error_count += 1
                        continue

                    # Validar DNI con 8 caracteres
                    if document_type == Patient.DocumentType.DNI and len(str(document_number)) != 8:
                        logger.warning(
                            f"Fila {row_number}: DNI inválido '{document_number}' tiene {len(str(document_number))} caracteres, debe tener 8 "
                            f"[{last_name}, {first_name}]"
                        )
                        skipped_count += 1
                        continue

                    # Mapear sexo
                    sex = sex_map.get(sex_excel)
                    if not sex:
                        logger.warning(
                            f"Fila {row_number}: Sexo inválido '{sex_excel}' "
                            f"[{last_name}, {first_name}, {document_type_excel} {document_number}] - Valores permitidos: F, M"
                        )
                        error_count += 1
                        continue

                    # Parsear fecha (formato: d/M/Y, ej: 17/10/2023)
                    # Si viene como string
                    if isinstance(birthdate_str, str):
                        birthdate = datetime.strptime(birthdate_str, "%d/%m/%Y").date()
                    else:
                        # Si viene como datetime de Excel
                        birthdate = birthdate_str.date() if hasattr(birthdate_str, "date") else birthdate_str

                    # Verificar si paciente ya existe (por document_type + document_number)
                    if Patient.objects.filter(
                        document_type=document_type, document_number=str(document_number)
                    ).exists():
                        logger.warning(
                            f"Fila {row_number}: Paciente duplicado - Ya existe con {document_type_excel} {document_number} "
                            f"[{last_name}, {first_name}]"
                        )
                        skipped_count += 1
                        continue

                    # Crear paciente
                    Patient.objects.create(
                        document_type=document_type,
                        document_number=str(document_number),
                        first_name=first_name,
                        last_name=last_name,
                        birthdate=birthdate,
                        sex=sex,
                        phone_number=str(phone_number),
                        email="",  # Optional
                        lead_source=None,  # Optional
                    )

                    created_count += 1

                except Exception as e:
                    logger.warning(
                        f"Fila {row_number}: Error al procesar - {str(e)} "
                        f"[{last_name}, {first_name}, {document_type_excel} {document_number}]"
                    )
                    logger.exception(f"Detalles del error en fila {row_number}")
                    error_count += 1
                    continue

            messages.success(
                request,
                f"Importación completada: {created_count} creados, {skipped_count} omitidos, {error_count} errores",
            )

        except Exception as e:
            logger.exception("Error al procesar el archivo de pacientes")
            messages.error(request, f"Error al procesar el archivo: {str(e)}")

        return redirect("patients_list")


class AdmissionView(LoginRequiredMixin, TemplateView):
    template_name = "patients/admission.html"
    login_url = reverse_lazy("login")


@require_GET
def search_patient_api(request):
    """API endpoint para buscar pacientes por tipo y número de documento o por query general"""
    if not request.user.is_authenticated:
        return JsonResponse({"error": "No autenticado"}, status=401)

    # Búsqueda nueva por query (nombre, apellido, documento)
    query = request.GET.get("query", "").strip()
    if query:
        if len(query) < 2:
            return JsonResponse({"patients": []})

        # Buscar por nombre, apellido o documento
        patients = Patient.objects.filter(
            models.Q(first_name__icontains=query)
            | models.Q(last_name__icontains=query)
            | models.Q(document_number__icontains=query)
        )[:10]

        patients_data = [
            {
                "id": patient.id,
                "first_name": patient.first_name,
                "last_name": patient.last_name,
                "document_type": patient.document_type,
                "document_number": patient.document_number,
            }
            for patient in patients
        ]

        return JsonResponse({"patients": patients_data})

    # Búsqueda antigua por tipo y número de documento (para compatibilidad)
    document_type = request.GET.get("document_type")
    document_number = request.GET.get("document_number")

    if not document_type or not document_number:
        return JsonResponse({"error": "Parámetros incompletos"}, status=400)

    try:
        patient = Patient.objects.get(document_type=document_type, document_number=document_number)
        return JsonResponse(
            {
                "found": True,
                "patient": {
                    "id": patient.id,
                    "document_type": patient.document_type,
                    "document_number": patient.document_number,
                    "first_name": patient.first_name,
                    "last_name": patient.last_name,
                },
            }
        )
    except Patient.DoesNotExist:
        return JsonResponse({"found": False, "message": "Paciente no encontrado"})


class LeadSourceListView(LoginRequiredMixin, ListView):
    model = LeadSource
    template_name = "patients/lead_source_list.html"
    context_object_name = "lead_sources"
    paginate_by = 20
    login_url = reverse_lazy("login")

    def get_queryset(self):
        return LeadSource.objects.all().order_by("name")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["breadcrumbs"] = [
            {"name": "Canales de Adquisición", "url": None},
        ]
        return context


class CreateLeadSourceView(LoginRequiredMixin, CreateView):
    model = LeadSource
    form_class = LeadSourceForm
    template_name = "patients/lead_source_create.html"
    success_url = reverse_lazy("lead_sources_list")
    login_url = reverse_lazy("login")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["breadcrumbs"] = [
            {"name": "Canales de Adquisición", "url": reverse_lazy("lead_sources_list")},
            {"name": "Crear Canal", "url": None},
        ]
        return context

    def form_valid(self, form):
        messages.success(self.request, "Canal de adquisición creado exitosamente")
        return super().form_valid(form)


class UpdateLeadSourceView(LoginRequiredMixin, UpdateView):
    model = LeadSource
    form_class = LeadSourceForm
    template_name = "patients/lead_source_update.html"
    success_url = reverse_lazy("lead_sources_list")
    login_url = reverse_lazy("login")
    context_object_name = "lead_source"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["breadcrumbs"] = [
            {"name": "Canales de Adquisición", "url": reverse_lazy("lead_sources_list")},
            {"name": f"Editar: {self.object.name}", "url": None},
        ]
        return context

    def form_valid(self, form):
        messages.success(self.request, "Canal de adquisición actualizado exitosamente")
        return super().form_valid(form)
