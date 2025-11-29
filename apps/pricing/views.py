from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import CreateView, ListView, UpdateView

from apps.exams.models import Exam
from apps.pricing.models import Coupon, PriceList, PriceListItem
from apps.referrals.models import Referral

from .services import PricingService


# PriceList Views
class PriceListListView(LoginRequiredMixin, ListView):
    model = PriceList
    template_name = "pricing/price_list_list.html"
    context_object_name = "price_lists"
    paginate_by = 10
    login_url = reverse_lazy("login")

    def get_queryset(self):
        return PriceList.objects.all().order_by("-created_at")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["breadcrumbs"] = [
            {"name": "Tarifarios", "url": None},
        ]
        return context


class PriceListCreateView(LoginRequiredMixin, CreateView):
    model = PriceList
    template_name = "pricing/price_list_create.html"
    fields = ["name", "is_active"]
    success_url = reverse_lazy("price_list_list")
    login_url = reverse_lazy("login")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["breadcrumbs"] = [
            {"name": "Tarifarios", "url": reverse_lazy("price_list_list")},
            {"name": "Crear Tarifario", "url": None},
        ]
        return context

    def form_valid(self, form):
        messages.success(self.request, f"Tarifario '{form.instance.name}' creado exitosamente")
        return super().form_valid(form)


class PriceListUpdateView(LoginRequiredMixin, UpdateView):
    model = PriceList
    template_name = "pricing/price_list_update.html"
    fields = ["name", "is_active"]
    success_url = reverse_lazy("price_list_list")
    login_url = reverse_lazy("login")
    context_object_name = "price_list"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["breadcrumbs"] = [
            {"name": "Tarifarios", "url": reverse_lazy("price_list_list")},
            {"name": f"Editar: {self.object.name}", "url": None},
        ]
        return context

    def form_valid(self, form):
        messages.success(self.request, f"Tarifario '{form.instance.name}' actualizado exitosamente")
        return super().form_valid(form)


class PriceListDetailView(LoginRequiredMixin, ListView):
    """Detail view showing all items in a price list"""

    model = PriceListItem
    template_name = "pricing/price_list_detail.html"
    context_object_name = "items"
    paginate_by = 20
    login_url = reverse_lazy("login")

    def get_queryset(self):
        self.price_list = get_object_or_404(PriceList, pk=self.kwargs["pk"])
        queryset = (
            PriceListItem.objects.filter(price_list=self.price_list).select_related("exam").order_by("exam__name")
        )

        # Filter by exam name if provided
        search = self.request.GET.get("search")
        if search:
            queryset = queryset.filter(exam__name__icontains=search)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["price_list"] = self.price_list
        context["search"] = self.request.GET.get("search", "")
        context["breadcrumbs"] = [
            {"name": "Tarifarios", "url": reverse_lazy("price_list_list")},
            {"name": self.price_list.name, "url": None},
        ]
        return context


class PriceListDownloadView(LoginRequiredMixin, View):
    """Download current price list (Excel with assigned exams and prices)"""

    login_url = reverse_lazy("login")

    def get(self, request, pk):
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        price_list = get_object_or_404(PriceList, pk=pk)

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tarifario"

        # Headers
        headers = ["Código", "Nombre del Examen", "Precio"]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Get price list items for this specific price list
        items = (
            PriceListItem.objects.filter(price_list=price_list)
            .select_related("exam")
            .order_by("exam__code", "exam__name")
        )

        for item in items:
            ws.append([item.exam.code or "", item.exam.name, float(item.price)])

        # Adjust column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 15

        # Create response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="tarifario_{price_list.name}.xlsx"'

        wb.save(response)
        return response


class PriceListDownloadTemplateView(LoginRequiredMixin, View):
    """Download template with all available exams for uploading"""

    login_url = reverse_lazy("login")

    def get(self, request, pk):
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        price_list = get_object_or_404(PriceList, pk=pk)

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plantilla Tarifario"

        # Headers
        headers = ["Código", "Nombre del Examen", "Precio"]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Get all exams
        exams = Exam.objects.all().order_by("code", "name")

        for exam in exams:
            ws.append([exam.code or "", exam.name, float(exam.price)])

        # Adjust column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 15

        # Create response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="plantilla_tarifario_{price_list.name}.xlsx"'

        wb.save(response)
        return response


class PriceListUploadView(LoginRequiredMixin, View):
    """Upload price list from Excel file"""

    login_url = reverse_lazy("login")
    template_name = "pricing/price_list_upload.html"

    def get(self, request, pk):
        price_list = get_object_or_404(PriceList, pk=pk)
        context = {
            "price_list": price_list,
            "breadcrumbs": [
                {"name": "Tarifarios", "url": reverse_lazy("price_list_list")},
                {"name": price_list.name, "url": reverse_lazy("price_list_detail", kwargs={"pk": pk})},
                {"name": "Subir Tarifario", "url": None},
            ],
        }
        return self.render_to_response(context)

    def post(self, request, pk):
        price_list = get_object_or_404(PriceList, pk=pk)

        if "file" not in request.FILES:
            messages.error(request, "No se seleccionó ningún archivo")
            return redirect("price_list_list")

        excel_file = request.FILES["file"]

        try:
            import openpyxl

            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            created_count = 0
            updated_count = 0
            error_count = 0

            # Skip header row
            for row in ws.iter_rows(min_row=2, values_only=True):
                code = row[0]
                price = row[2]

                if not code or not price:
                    continue

                try:
                    # Find exam by code
                    exam = Exam.objects.filter(code=code).first()

                    if not exam:
                        error_count += 1
                        continue

                    # Create or update price list item
                    item, created = PriceListItem.objects.update_or_create(
                        price_list=price_list, exam=exam, defaults={"price": price}
                    )

                    if created:
                        created_count += 1
                    else:
                        updated_count += 1

                except Exception:
                    error_count += 1
                    continue

            messages.success(
                request,
                f"Tarifario cargado: {created_count} nuevos, {updated_count} actualizados, {error_count} errores",
            )

        except Exception as e:
            messages.error(request, f"Error al procesar el archivo: {str(e)}")

        return redirect("price_list_list")

    def render_to_response(self, context):
        from django.shortcuts import render

        return render(self.request, self.template_name, context)


# Coupon Views
class CouponListView(LoginRequiredMixin, ListView):
    model = Coupon
    template_name = "pricing/coupon_list.html"
    context_object_name = "coupons"
    paginate_by = 10
    login_url = reverse_lazy("login")

    def get_queryset(self):
        return Coupon.objects.select_related("price_list").all().order_by("-created_at")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["breadcrumbs"] = [
            {"name": "Cupones", "url": None},
        ]
        return context


class CouponCreateView(LoginRequiredMixin, CreateView):
    model = Coupon
    template_name = "pricing/coupon_create.html"
    fields = ["code", "price_list", "expiration_date", "is_active"]
    success_url = reverse_lazy("coupon_list")
    login_url = reverse_lazy("login")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["price_lists"] = PriceList.objects.filter(is_active=True).order_by("name")
        context["breadcrumbs"] = [
            {"name": "Cupones", "url": reverse_lazy("coupon_list")},
            {"name": "Crear Cupón", "url": None},
        ]
        return context

    def form_valid(self, form):
        messages.success(self.request, f"Cupón '{form.instance.code}' creado exitosamente")
        return super().form_valid(form)


class CouponUpdateView(LoginRequiredMixin, UpdateView):
    model = Coupon
    template_name = "pricing/coupon_update.html"
    fields = ["code", "price_list", "expiration_date", "is_active"]
    success_url = reverse_lazy("coupon_list")
    login_url = reverse_lazy("login")
    context_object_name = "coupon"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["price_lists"] = PriceList.objects.filter(is_active=True).order_by("name")
        context["breadcrumbs"] = [
            {"name": "Cupones", "url": reverse_lazy("coupon_list")},
            {"name": f"Editar: {self.object.code}", "url": None},
        ]
        return context

    def form_valid(self, form):
        messages.success(self.request, f"Cupón '{form.instance.code}' actualizado exitosamente")
        return super().form_valid(form)


# Referral Views
class ReferralListView(LoginRequiredMixin, ListView):
    model = Referral
    template_name = "pricing/referral_list.html"
    context_object_name = "referrals"
    paginate_by = 10
    login_url = reverse_lazy("login")

    def get_queryset(self):
        return Referral.objects.select_related("price_list").all().order_by("-created_at")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["breadcrumbs"] = [
            {"name": "Referidos", "url": None},
        ]
        return context


class ReferralCreateView(LoginRequiredMixin, CreateView):
    model = Referral
    template_name = "pricing/referral_create.html"
    fields = ["business_name", "document_number", "phone_number", "email", "address", "price_list", "is_active"]
    success_url = reverse_lazy("referral_list")
    login_url = reverse_lazy("login")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["price_lists"] = PriceList.objects.filter(is_active=True).order_by("name")
        context["breadcrumbs"] = [
            {"name": "Referidos", "url": reverse_lazy("referral_list")},
            {"name": "Crear Referido", "url": None},
        ]
        return context

    def form_valid(self, form):
        messages.success(self.request, f"Referido '{form.instance.business_name}' creado exitosamente")
        return super().form_valid(form)


class ReferralUpdateView(LoginRequiredMixin, UpdateView):
    model = Referral
    template_name = "pricing/referral_update.html"
    fields = ["business_name", "document_number", "phone_number", "email", "address", "price_list", "is_active"]
    success_url = reverse_lazy("referral_list")
    login_url = reverse_lazy("login")
    context_object_name = "referral"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["price_lists"] = PriceList.objects.filter(is_active=True).order_by("name")
        context["breadcrumbs"] = [
            {"name": "Referidos", "url": reverse_lazy("referral_list")},
            {"name": f"Editar: {self.object.business_name}", "url": None},
        ]
        return context

    def form_valid(self, form):
        messages.success(self.request, f"Referido '{form.instance.business_name}' actualizado exitosamente")
        return super().form_valid(form)


# API Endpoints
def get_exam_price_api(request):
    """
    API endpoint para obtener el precio de un examen considerando cupón, tarifario del referido o precio base.

    Query params:
        - exam_id: ID del examen (requerido)
        - referral_id: ID del referido (opcional)
        - coupon_code: Código del cupón (opcional)

    Returns:
        JSON con:
            - price: precio del examen
            - source: 'coupon', 'price_list' o 'base'
            - price_list_id: ID del tarifario (si aplica)
            - coupon_code: código del cupón (si aplica)
    """
    if not request.user.is_authenticated:
        return JsonResponse({"error": "No autenticado"}, status=401)

    exam_id = request.GET.get("exam_id")
    referral_id = request.GET.get("referral_id")
    coupon_code = request.GET.get("coupon_code", "").strip()

    if not exam_id:
        return JsonResponse({"error": "exam_id es requerido"}, status=400)

    try:
        # Convertir a int
        exam_id = int(exam_id)
        referral_id = int(referral_id) if referral_id else None

        # Usar el servicio para obtener el precio
        result = PricingService.get_exam_price(exam_id, referral_id, coupon_code or None)

        return JsonResponse(result)

    except ValueError:
        return JsonResponse({"error": "IDs inválidos"}, status=400)
    except Exam.DoesNotExist:
        return JsonResponse({"error": "Examen no encontrado"}, status=404)
    except Referral.DoesNotExist:
        return JsonResponse({"error": "Referido no encontrado"}, status=404)
    except Exception as e:
        return JsonResponse({"error": f"Error al obtener precio: {str(e)}"}, status=500)


def validate_coupon_api(request):
    """
    API endpoint para validar un código de cupón.

    Query params:
        - coupon_code: Código del cupón (requerido)

    Returns:
        JSON con:
            - valid: True/False
            - coupon: datos del cupón (si es válido)
            - error: mensaje de error (si no es válido)
    """
    if not request.user.is_authenticated:
        return JsonResponse({"error": "No autenticado"}, status=401)

    coupon_code = request.GET.get("coupon_code", "").strip()

    if not coupon_code:
        return JsonResponse({"error": "Código de cupón requerido"}, status=400)

    try:
        result = PricingService.validate_coupon(coupon_code)
        return JsonResponse(result)

    except Exception as e:
        return JsonResponse({"error": f"Error al validar cupón: {str(e)}"}, status=500)
