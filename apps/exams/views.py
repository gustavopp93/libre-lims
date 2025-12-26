import logging

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import CreateView, ListView, UpdateView

from apps.exams.forms import (
    ExamCategoryForm,
    ExamCategoryUpdateForm,
    ExamComponentFormSet,
    ExamForm,
    ExamUpdateForm,
)
from apps.exams.models import Exam, ExamCategory

logger = logging.getLogger(__name__)


class ExamsListView(LoginRequiredMixin, ListView):
    model = Exam
    template_name = "exams/exam_list.html"
    context_object_name = "exams"
    paginate_by = 20
    login_url = reverse_lazy("login")

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filtrar por nombre
        name = self.request.GET.get("name")
        if name:
            queryset = queryset.filter(name__icontains=name)

        return queryset.order_by("-created_at")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["name"] = self.request.GET.get("name", "")
        context["breadcrumbs"] = [
            {"name": "Exámenes", "url": None},
        ]
        return context


class CreateExamView(LoginRequiredMixin, CreateView):
    model = Exam
    form_class = ExamForm
    template_name = "exams/exam_create.html"
    success_url = reverse_lazy("exams_list")
    login_url = reverse_lazy("login")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["breadcrumbs"] = [
            {"name": "Exámenes", "url": reverse_lazy("exams_list")},
            {"name": "Crear Examen", "url": None},
        ]
        if self.request.POST:
            # Usar POST limpio si existe (cuando hubo error de validación)
            post_data = getattr(self, "_cleaned_post", self.request.POST)
            context["component_formset"] = ExamComponentFormSet(post_data)
        else:
            context["component_formset"] = ExamComponentFormSet()

        return context

    def form_valid(self, form):
        context = self.get_context_data()
        component_formset = context["component_formset"]

        # Validar que si tiene componentes, debe haber al menos uno
        has_components = form.cleaned_data.get("has_components", False)
        if has_components:
            # Validar el formset primero
            if not component_formset.is_valid():
                return self.form_invalid(form)

            # Contar componentes válidos (no marcados para eliminar)
            valid_components = sum(
                1
                for form_item in component_formset
                if hasattr(form_item, "cleaned_data")
                and form_item.cleaned_data
                and not form_item.cleaned_data.get("DELETE", False)
            )
            if valid_components == 0:
                form.add_error("has_components", "Debe agregar al menos un componente al examen.")

                # Limpiar los DELETE del POST para que los componentes vuelvan a aparecer
                mutable_post = self.request.POST.copy()
                for i in range(len(component_formset)):
                    delete_field = f"component_items-{i}-DELETE"
                    if delete_field in mutable_post:
                        del mutable_post[delete_field]

                # Almacenar el POST limpio para que get_context_data() lo use
                self._cleaned_post = mutable_post

                return self.form_invalid(form)

        with transaction.atomic():
            # Generar código automáticamente
            last_exam = Exam.objects.order_by("-code").first()
            if last_exam and last_exam.code and last_exam.code.startswith("EX"):
                try:
                    last_number = int(last_exam.code[2:])
                    new_number = last_number + 1
                except ValueError:
                    new_number = 1
            else:
                new_number = 1

            form.instance.code = f"EX{new_number:05d}"

            self.object = form.save()

            if component_formset.is_valid():
                component_formset.instance = self.object
                component_formset.save()
            else:
                return self.form_invalid(form)

        messages.success(self.request, "Examen creado exitosamente")
        return super().form_valid(form)


class UpdateExamView(LoginRequiredMixin, UpdateView):
    model = Exam
    form_class = ExamUpdateForm
    template_name = "exams/exam_update.html"
    success_url = reverse_lazy("exams_list")
    login_url = reverse_lazy("login")
    context_object_name = "exam"

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Deshabilitar el checkbox de componentes si:
        # 1. El examen ya tiene componentes guardados
        # 2. El examen es usado como componente de otro examen
        has_saved_components = self.object.component_items.exists()
        is_used_as_component = self.object.parent_exams.exists()

        if has_saved_components or is_used_as_component:
            form.fields["has_components"].widget.attrs["disabled"] = "disabled"
            form.fields["has_components"].widget.attrs["title"] = (
                "No se puede modificar porque el examen tiene componentes guardados"
                if has_saved_components
                else "No se puede modificar porque este examen es usado como componente de otro examen"
            )

        return form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["breadcrumbs"] = [
            {"name": "Exámenes", "url": reverse_lazy("exams_list")},
            {"name": f"Editar: {self.object.name}", "url": None},
        ]
        if self.request.POST:
            # Usar POST limpio si existe (cuando hubo error de validación)
            post_data = getattr(self, "_cleaned_post", self.request.POST)
            context["component_formset"] = ExamComponentFormSet(post_data, instance=self.object)
        else:
            # Ordenar los componentes por el campo 'order'
            context["component_formset"] = ExamComponentFormSet(
                instance=self.object, queryset=self.object.component_items.order_by("order")
            )

        return context

    def form_valid(self, form):
        context = self.get_context_data()
        component_formset = context["component_formset"]

        # Si el checkbox está deshabilitado, restaurar el valor original
        has_saved_components = self.object.component_items.exists()
        is_used_as_component = self.object.parent_exams.exists()
        if has_saved_components or is_used_as_component:
            form.instance.has_components = self.object.has_components

        # Validar que si tiene componentes, debe haber al menos uno
        has_components = form.instance.has_components
        if has_components:
            # Validar el formset primero
            if not component_formset.is_valid():
                return self.form_invalid(form)

            # Contar componentes válidos (no marcados para eliminar)
            valid_components = sum(
                1
                for form_item in component_formset
                if hasattr(form_item, "cleaned_data")
                and form_item.cleaned_data
                and not form_item.cleaned_data.get("DELETE", False)
            )
            if valid_components == 0:
                form.add_error("has_components", "Debe agregar al menos un componente al examen.")

                # Limpiar los DELETE del POST para que los componentes vuelvan a aparecer
                mutable_post = self.request.POST.copy()
                for i in range(len(component_formset)):
                    delete_field = f"component_items-{i}-DELETE"
                    if delete_field in mutable_post:
                        del mutable_post[delete_field]

                # Almacenar el POST limpio para que get_context_data() lo use
                self._cleaned_post = mutable_post

                return self.form_invalid(form)

        with transaction.atomic():
            self.object = form.save()

            if component_formset.is_valid():
                component_formset.instance = self.object
                component_formset.save()
            else:
                return self.form_invalid(form)

        messages.success(self.request, "Examen actualizado exitosamente")
        return super().form_valid(form)


@login_required
def search_exams_api(request):
    """API endpoint para buscar exámenes por nombre"""
    name = request.GET.get("name", "")
    parent_exam_id = request.GET.get("parent_exam_id", None)

    if not name:
        return JsonResponse({"exams": []})

    exams = Exam.objects.filter(name__icontains=name)

    # Excluir el examen padre (validar que sea un entero válido)
    if parent_exam_id:
        try:
            parent_exam_id = int(parent_exam_id)
            exams = exams.exclude(id=parent_exam_id)
        except (ValueError, TypeError):
            # Si no es un entero válido, ignorar el filtro
            pass

    exams = exams[:10]  # Limitar a 10 resultados

    exams_data = [
        {
            "id": exam.id,
            "name": exam.name,
        }
        for exam in exams
    ]

    return JsonResponse({"exams": exams_data})


# ==================== EXAM CATEGORIES ====================


class ExamCategoriesListView(LoginRequiredMixin, ListView):
    model = ExamCategory
    template_name = "exams/category_list.html"
    context_object_name = "categories"
    paginate_by = 20
    login_url = reverse_lazy("login")

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filtrar por nombre
        name = self.request.GET.get("name")
        if name:
            queryset = queryset.filter(name__icontains=name)

        return queryset.order_by("name")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["name"] = self.request.GET.get("name", "")
        context["breadcrumbs"] = [
            {"name": "Categorías de Exámenes", "url": None},
        ]
        return context


class CreateExamCategoryView(LoginRequiredMixin, CreateView):
    model = ExamCategory
    form_class = ExamCategoryForm
    template_name = "exams/category_create.html"
    success_url = reverse_lazy("exam_categories_list")
    login_url = reverse_lazy("login")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["breadcrumbs"] = [
            {"name": "Categorías de Exámenes", "url": reverse_lazy("exam_categories_list")},
            {"name": "Crear Categoría", "url": None},
        ]
        return context

    def form_valid(self, form):
        # Generar código automáticamente
        last_category = ExamCategory.objects.order_by("-code").first()
        if last_category and last_category.code.startswith("CA"):
            try:
                last_number = int(last_category.code[2:])
                new_number = last_number + 1
            except ValueError:
                new_number = 1
        else:
            new_number = 1

        form.instance.code = f"CA{new_number:03d}"

        messages.success(self.request, "Categoría creada exitosamente")
        return super().form_valid(form)


class UpdateExamCategoryView(LoginRequiredMixin, UpdateView):
    model = ExamCategory
    form_class = ExamCategoryUpdateForm
    template_name = "exams/category_update.html"
    success_url = reverse_lazy("exam_categories_list")
    login_url = reverse_lazy("login")
    context_object_name = "category"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["breadcrumbs"] = [
            {"name": "Categorías de Exámenes", "url": reverse_lazy("exam_categories_list")},
            {"name": f"Editar: {self.object.name}", "url": None},
        ]
        return context

    def form_valid(self, form):
        messages.success(self.request, "Categoría actualizada exitosamente")
        return super().form_valid(form)


class ExamsDownloadTemplateView(LoginRequiredMixin, View):
    """Download template Excel for exams import"""

    login_url = reverse_lazy("login")

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plantilla Exámenes"

        # Headers: Nombre del Examen, Precio, Código de Categoría
        headers = ["Nombre del Examen", "Precio", "Código de Categoría"]
        ws.append(headers)

        # Estilizar headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Fila de ejemplo
        ws.append(["Hemograma Completo", 25.50, "CA001"])

        # Ajustar anchos de columna
        ws.column_dimensions["A"].width = 40  # Nombre del Examen
        ws.column_dimensions["B"].width = 15  # Precio
        ws.column_dimensions["C"].width = 20  # Código de Categoría

        # Create response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="plantilla_examenes.xlsx"'

        wb.save(response)
        return response


class ExamsUploadView(LoginRequiredMixin, View):
    """Upload exams from Excel file"""

    login_url = reverse_lazy("login")
    template_name = "exams/exams_upload.html"

    def get(self, request):
        context = {
            "breadcrumbs": [
                {"name": "Exámenes", "url": reverse_lazy("exams_list")},
                {"name": "Importar Exámenes", "url": None},
            ],
        }
        return render(request, self.template_name, context)

    def post(self, request):
        if "file" not in request.FILES:
            messages.error(request, "No se seleccionó ningún archivo")
            return redirect("exams_list")

        excel_file = request.FILES["file"]

        try:
            import openpyxl

            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            created_count = 0
            skipped_count = 0
            error_count = 0

            # Iterar desde fila 2 (skip header)
            for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Columnas: Nombre del Examen, Precio, Código de Categoría
                exam_name = row[0]
                price = row[1]
                category_code = row[2]

                # Skip si faltan datos críticos
                if not all([exam_name, price, category_code]):
                    missing_fields = []
                    if not exam_name:
                        missing_fields.append("Nombre del Examen")
                    if not price:
                        missing_fields.append("Precio")
                    if not category_code:
                        missing_fields.append("Código de Categoría")

                    logger.warning(
                        f"Fila {row_number}: Datos incompletos - Faltan campos: {', '.join(missing_fields)} "
                        f"[{exam_name or 'N/A'}, {price or 'N/A'}, {category_code or 'N/A'}]"
                    )
                    error_count += 1
                    continue

                try:
                    # Aplicar strip() para limpiar espacios en blanco
                    exam_name = str(exam_name).strip()
                    category_code = str(category_code).strip()

                    # Verificar si el examen ya existe (case insensitive)
                    if Exam.objects.filter(name__iexact=exam_name).exists():
                        logger.warning(
                            f"Fila {row_number}: Examen duplicado - Ya existe un examen con el nombre '{exam_name}'"
                        )
                        skipped_count += 1
                        continue

                    # Buscar categoría por código
                    try:
                        category = ExamCategory.objects.get(code=category_code)
                    except ExamCategory.DoesNotExist:
                        logger.warning(
                            f"Fila {row_number}: Categoría no encontrada con código '{category_code}' "
                            f"[{exam_name}, {price}]"
                        )
                        error_count += 1
                        continue

                    # Validar precio
                    try:
                        price = float(price)
                        if price < 0:
                            logger.warning(
                                f"Fila {row_number}: Precio inválido '{price}' debe ser mayor o igual a 0 "
                                f"[{exam_name}, {category_code}]"
                            )
                            error_count += 1
                            continue
                    except (ValueError, TypeError):
                        logger.warning(
                            f"Fila {row_number}: Precio inválido '{price}' debe ser un número "
                            f"[{exam_name}, {category_code}]"
                        )
                        error_count += 1
                        continue

                    # Generar código automáticamente
                    last_exam = Exam.objects.order_by("-code").first()
                    if last_exam and last_exam.code and last_exam.code.startswith("EX"):
                        try:
                            last_number = int(last_exam.code[2:])
                            new_number = last_number + 1
                        except ValueError:
                            new_number = 1
                    else:
                        new_number = 1

                    exam_code = f"EX{new_number:05d}"

                    # Crear examen
                    Exam.objects.create(
                        code=exam_code,
                        name=exam_name,
                        price=price,
                        category=category,
                        has_components=False,
                    )

                    created_count += 1

                except Exception as e:
                    logger.warning(
                        f"Fila {row_number}: Error al procesar - {str(e)} " f"[{exam_name}, {price}, {category_code}]"
                    )
                    logger.exception(f"Detalles del error en fila {row_number}")
                    error_count += 1
                    continue

            messages.success(
                request,
                f"Importación completada: {created_count} creados, {skipped_count} omitidos, {error_count} errores",
            )

        except Exception as e:
            logger.exception("Error al procesar el archivo de exámenes")
            messages.error(request, f"Error al procesar el archivo: {str(e)}")

        return redirect("exams_list")
